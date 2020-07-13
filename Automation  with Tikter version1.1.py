
# for automation of printing as it is windows dialog
import autoit
# for printing as chrome printing is in json format
import json
# For GUI
import tkinter as tk
from tkinter import filedialog        
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook, cell
from time import sleep

# #####################################################
# It Clicks Save as The Print Prompt Opens Enter
options = Options()
appState = {
    "recentDestinations": [
        {
            "id": "Save as PDF",
            "origin": "local"
        }
    ],
    "selectedDestinationId": "Save as PDF",
    "version": 2
}
profile = {'printing.print_preview_sticky_settings.appState': json.dumps(appState)}
# profile = {'printing.print_preview_sticky_settings.appState':json.dumps(appState),'savefile.default_directory':downloadPath}
options.add_experimental_option('prefs', profile)
options.add_argument('--kiosk-printing')
CHROMEDRIVER_PATH = 'D:/Pythonfile/chromedriver.exe'
# ################################################
root = tk.Tk()
file = filedialog.askopenfilename(initialdir = "/", 
                                          title = "Select a File", 
                                          filetypes = (("Excel files", 
                                                        "*.xlsx*"), 
                                                       ("all files", 
                                                        "*.*")))
root.destroy()

# Number of rows to execute in a file 
total_num_rows = int(input('Enter The Number of WorkOrders(rows) in your files ==> '))
# Add 1 because we start from 2
total_num_rows = total_num_rows + 1

# Column to show comments 6 means F column
comments = 6
#total_num_rows = 20

driver = webdriver.Chrome('chromedriver', options=options,)

# For login

driver.get('https://104i-sgapp.teleows.com/app/portal/loadPortal.action')
username = driver.find_element_by_id("usernameInput")
username.clear()
username.send_keys("")

password = driver.find_element_by_id("password")
password.clear()
password.send_keys("")

driver.find_element_by_id("btn_submit").click()

# ############## Login ENd ##########################
wb = load_workbook(file)
wbsheet = wb["Sheet1"]

# ############### For Page Scroll ####################
def scroll_down_page(speed=8):
        current_scroll_position, new_height = 0, 1
        while current_scroll_position <= new_height:
            current_scroll_position += speed
            driver.execute_script("window.scrollTo(0, {});".format(current_scroll_position))
            new_height = driver.execute_script("return document.body.scrollHeight")
# ################ Loop ####################################
    
for x in range(2,total_num_rows):
	Task_id = wbsheet.cell(row=x, column=1).value
	wbsheet.cell(row=x,column=6).value="Processing"

	driver.get("https://104i-sgapp.teleows.com/app/104i/spl/c_mission_control_service/mission_task_work_grid.spl")

	SelectPm = driver.find_element_by_id('task_type')
	SelectPm.clear()
	SelectPm.send_keys("PM")
	start = driver.find_element_by_id('start_time_input')
	start.clear()

	end = driver.find_element_by_id('end_time_input')
	end.clear()
	driver.find_element_by_id('ToolbarExtender1').click()
	# self.driver.find_element_by_id('search').click()

	taskId = driver.find_element_by_name('task_id')
	taskId.clear()
	
	taskId.send_keys(Task_id)

	driver.find_element_by_id('search').click()
	sleep(12)
	try:
		############click on task id .... link ###################
		driver.find_element_by_xpath('//*[@id="ext-gen163"]/a').click()
	except NoSuchElementException:
		wbsheet.cell(row=x, column=comments).value =  'No Link Found!'
		wb.save(file)
	else:
		# driver.find_element_by_css_selector("#ext-gen109 > a:nth-child(1)").click()
		driver.switch_to.window(driver.window_handles[-1])
		sleep(10)
		try:
			driver.find_element_by_xpath('//*[@id="checklist"]').click()
		except NoSuchElementException:
			wbsheet.cell(row=x, column=comments).value =  'No checklist Found!'
			wb.save(file)
		else:
			driver.switch_to.window(driver.window_handles[-1])
			sleep(5)
			siteId = wbsheet.cell(row=x, column=2).value
			######  slow 2 , fast 20 ############
			######  you can increase speed from here  ############
			scroll_down_page(3)
			scroll_down_page(4)
			scroll_down_page(12)
			driver.execute_script('window.print();')
			sleep(5)
			autoit.control_send("Save Print Output As", "Edit1", str(siteId))
			autoit.control_send("Save Print Output As", "Button2", "Save")
			autoit.send("Enter")
			wbsheet.cell(row=x, column=comments).value =  'Done'
			wb.save(file)
